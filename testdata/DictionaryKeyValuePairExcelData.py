import openpyxl

book = openpyxl.load_workbook("E:\\PythonTesting\\PythonTestData.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
Dict = {}
sheet.cell(row=2, column=3).value ="sinha"
print(sheet.cell(row=2, column=3).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A3'].value)
for i in range(1, sheet.max_row+1): # for row
    if sheet.cell(row =i, column = 1).value == "Testcase2": # for specific row

        for j in range (1, sheet.max_column+1): # to get column values
            Dict[sheet.cell(row =1, column =j).value]= sheet.cell(row=i, column=j).value

print(Dict)

